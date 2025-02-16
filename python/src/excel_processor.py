import os
from openpyxl import load_workbook
from datetime import datetime
import json
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd

class HomeworkData:
    """Class to represent the decoded JSON data structure."""
    def __init__(self, json_data: dict):
        self.date = datetime.strptime(json_data['metadata']['date']['raw'], '%Y-%m-%d')
        self.attendance = json_data['attendance']
        self.homework = json_data['homework']
        self.previous_homework = json_data['previousHomework']

class ExcelProcessor:
    """
    Excel processor that handles formula-based Excel files correctly,
    regardless of which application last saved them.
    """
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        # We'll maintain two workbooks:
        # One for reading values (data_only=True)
        # One for preserving formulas (data_only=False)
        self.formula_wb = None
        self.data_wb = None
    
    def __enter__(self):
        """Load both versions of the workbook when entering the context."""
        self.formula_wb = load_workbook(self.filepath, data_only=False)
        self.data_wb = load_workbook(self.filepath, data_only=True)
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Ensure both workbooks are closed on exit."""
        if self.formula_wb:
            self.formula_wb.close()
        if self.data_wb:
            self.data_wb.close()

    def _get_cell_value(self, worksheet, row: int, col: int) -> Any:
        """
        Get the actual value of a cell, handling both direct values and formulas.
        
        Args:
            worksheet: The worksheet from data_wb (with data_only=True)
            row: Row number (1-based)
            col: Column number (1-based)
            
        Returns:
            The calculated value of the cell
        """
        cell = worksheet.cell(row=row, column=col)
        return cell.value

    def find_header_and_date_row(self, sheet_name: str, target_date: datetime) -> Tuple[Optional[int], Optional[int]]:
        """
        Find both the header row and the target date row.
        
        This method uses the data_only workbook to read actual calculated values,
        making it reliable regardless of formula calculation status.
        
        Args:
            sheet_name: Name of the worksheet
            target_date: The date we're looking for
            
        Returns:
            Tuple of (header_row, date_row) numbers
        """
        # Get the worksheet from our data_only workbook
        worksheet = self.data_wb[sheet_name]
        
        # First find the header row
        header_row = None
        for row in range(1, worksheet.max_row + 1):
            value = self._get_cell_value(worksheet, row, 1)
            if value and isinstance(value, str) and "التاريخ: السبت" in value.strip():
                header_row = row
                break
        
        if not header_row:
            return None, None
            
        # Now look for the target date in subsequent rows
        target_date = target_date.date()  # Convert to date for comparison
        date_row = None
        
        for row in range(header_row + 1, worksheet.max_row + 1):
            value = self._get_cell_value(worksheet, row, 1)
            
            # Skip empty cells
            if not value:
                continue
                
            # Handle different date formats
            try:
                if isinstance(value, datetime):
                    cell_date = value.date()
                else:
                    # Try parsing as string
                    cell_date = pd.to_datetime(value).date()
                
                if cell_date == target_date:
                    date_row = row
                    break
            except (ValueError, TypeError):
                continue
        
        return header_row, date_row

    def find_homework_columns(self, sheet_name: str, header_row: int) -> Tuple[List[str], int]:
        """
        Find homework type columns and their repetition point.
        
        Uses the formula workbook to read the actual column headers,
        ensuring we get the exact text as it appears in the file.
        """
        worksheet = self.formula_wb[sheet_name]
        homework_types = []
        repetition_start = None
        seen_types = set()
        
        for col in range(3, worksheet.max_column + 1):
            value = worksheet.cell(row=header_row, column=col).value
            if not value:
                continue
                
            cell_text = str(value).strip()
            
            if cell_text in seen_types:
                repetition_start = col
                break
                
            if any(hw_type in cell_text for hw_type in ["مراجعة بعيدة", "مراجعة قريبة", "حفظ"]):
                homework_types.append(cell_text)
                seen_types.add(cell_text)
        
        return homework_types, repetition_start

    def update_student_worksheet(self, student_name: str, homework_data: HomeworkData) -> Dict[str, Any]:
        """Update a single student's worksheet with new data."""
        try:
            # Get both versions of the worksheet
            formula_ws = self.formula_wb[student_name]
            data_ws = self.data_wb[student_name]
            
            # Find our target rows using the data worksheet
            header_row, date_row = self.find_header_and_date_row(
                student_name,
                homework_data.date
            )
            
            if not header_row or not date_row:
                return {
                    'success': False,
                    'error': "Could not find header row or target date"
                }
            
            # Find homework columns using the formula worksheet
            homework_types, repetition_col = self.find_homework_columns(
                student_name,
                header_row
            )
            
            if not repetition_col:
                return {
                    'success': False,
                    'error': "Could not find homework type repetition"
                }
            
            # Make updates to the formula worksheet
            # Update attendance
            attendance_value = "حاضر" if homework_data.attendance[student_name]['present'] else "غائب"
            formula_ws.cell(row=date_row, column=2, value=attendance_value)
            
            # Update previous homework grades
            current_col = 3
            for hw_type in homework_types:
                if hw_type in homework_data.previous_homework:
                    grade = homework_data.previous_homework[hw_type].get(student_name, '')
                    if not grade:
                        continue
                    grade = float(grade)
                    formula_ws.cell(row=date_row, column=current_col, value=grade)
                current_col += 1
            
            # Update new homework assignments
            current_col = repetition_col
            for hw_type in homework_types:
                content = ''
                for assignment in homework_data.homework['assignments']:
                    is_assignment_relevant = not(assignment['assignedStudents']) or student_name in assignment['assignedStudents']
                    if (assignment['type'] == hw_type and is_assignment_relevant):
                        content = assignment['content']
                        break
                formula_ws.cell(row=date_row, column=current_col, value=content)
                current_col += 1
            
            return {'success': True, 'error': None}
            
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def process_workbook(self, homework_data: HomeworkData) -> Dict[str, Any]:
        """Process the entire workbook."""
        results = {}
        
        for student_name in homework_data.attendance.keys():
            if student_name in self.formula_wb.sheetnames:
                results[student_name] = self.update_student_worksheet(
                    student_name,
                    homework_data
                )
            else:
                results[student_name] = {
                    'success': False,
                    'error': f"Worksheet not found for student: {student_name}"
                }
        
        return results
    
    def save(self, output_path: Optional[str] = None):
        """
        Save the workbook with formulas preserved.
        Only saves the formula workbook as it contains all our updates.
        """
        save_path = output_path or self.filepath
        self.formula_wb.save(save_path)
        os.startfile(save_path)

def process_excel_file(
    excel_path: str,
    json_data: dict,
    output_path: Optional[str] = None
) -> Dict[str, Any]:
    """Process an Excel file with the provided JSON data."""
    homework_data = HomeworkData(json_data)
    
    with ExcelProcessor(excel_path) as processor:
        results = processor.process_workbook(homework_data)
        processor.save(output_path)
        
    return results