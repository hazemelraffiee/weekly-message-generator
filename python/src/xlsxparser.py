import pandas as pd
from openpyxl import load_workbook
import logging

# Configure logging to handle Arabic characters
logging.basicConfig(
    filename='update_log.txt', 
    level=logging.INFO, 
    format='%(asctime)s - %(message)s', 
    encoding='utf-8'
)

def update_student_records(file_path, update_data, output_file):
    """
    Reads an Excel file, updates the specified date row in each student sheet with the given data,
    and saves the modified file while keeping original formatting. Logs the process details.
    
    :param file_path: Path to the Excel file
    :param update_data: Dictionary containing metadata, attendance, and homework details
    :param output_file: The output file path for saving updates
    """
    date = update_data['metadata']['date']['raw']
    attendance_data = update_data['attendance']
    previous_homework_data = update_data['previousHomework']
    
    # Load the workbook to preserve formatting
    wb = load_workbook(file_path)
    
    expected_students = set(attendance_data.keys())
    found_students = set(wb.sheetnames)
    missing_sheets = expected_students - found_students
    extra_sheets = found_students - expected_students
    
    if missing_sheets:
        logging.warning(f"لم يتم العثور على أوراق الطلاب المتوقعة: {', '.join(missing_sheets)}")
    if extra_sheets:
        logging.info(f"تم العثور على أوراق طلاب إضافية: {', '.join(extra_sheets)}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logging.info(f"معالجة ورقة الطالب: {sheet_name}")
        
        # Find the row index where the date matches
        date_found = False
        row_index = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == date:
                    row_index = cell.row
                    date_found = True
                    break
        
        if not date_found:
            logging.warning(f"لم يتم العثور على التاريخ {date} في ورقة الطالب: {sheet_name}")
            continue
        
        # Update attendance
        student_name = sheet_name
        if student_name in attendance_data:
            ws.cell(row=row_index, column=2, value='حاضر' if attendance_data[student_name]['present'] else 'غائب')
            ws.cell(row=row_index, column=3, value=attendance_data[student_name]['lateMinutes'])
            logging.info(f"تم تحديث الحضور للطالب {student_name}")
        else:
            logging.warning(f"تم العثور على الطالب {student_name} في ورقة العمل ولكنه غير موجود في البيانات المقدمة")
        
        # Update previous homework scores
        for hw_type, students in previous_homework_data.items():
            if student_name in students:
                column_index = 4 if hw_type == 'حفظ' else 5
                ws.cell(row=row_index, column=column_index, value=students[student_name])
                logging.info(f"تم تحديث درجة {hw_type} للطالب {student_name}")
    
    # Save the workbook with formatting preserved
    wb.save(output_file)
    logging.info(f"تم حفظ الملف المحدث باسم {output_file}")
    print(f"تم حفظ الملف المحدث باسم {output_file}. تحقق من update_log.txt للحصول على التفاصيل.")

decoded_data = {
  "metadata": {
    "schoolName": "عمر بن الخطاب - حلقات يوم السبت",
    "className": "الفوج الثاني",
    "date": {
      "raw": "2025-01-29",
      "formatted": "الأربعاء، ٣٠  رجب  ١٤٤٦ هـ الموافق الأربعاء، 29  يناير  2025 م"
    },
    "header_class_name": "الفوج الثاني",
    "header_date": "الأربعاء، ٣٠  رجب  ١٤٤٦ هـ الموافق الأربعاء، 29  يناير  2025 م"
  },
  "attendance": {
    "زياد بالقاسمي": {
      "present": True,
      "lateMinutes": ""
    },
    "أحمد الدقاق": {
      "present": True,
      "lateMinutes": ""
    },
    "عبد الرحمن الشايب": {
      "present": True,
      "lateMinutes": ""
    },
    "أنس أبو الفتوح": {
      "present": True,
      "lateMinutes": ""
    },
    "محمد عودة": {
      "present": True,
      "lateMinutes": ""
    },
    "عمر سند": {
      "present": True,
      "lateMinutes": ""
    },
    "زيد حطاب": {
      "present": True,
      "lateMinutes": 15
    }
  },
  "homework": {
    "assignments": [
      {
        "type": "حفظ",
        "content": "الحاقة",
        "assignedStudents": []
      },
      {
        "type": "مراجعة قريبة",
        "content": "الغاشية",
        "assignedStudents": [
          "أنس أبو الفتوح"
        ]
      }
    ]
  },
  "previousHomework": {
    "حفظ": {
      "زياد بالقاسمي": "1",
      "أحمد الدقاق": "1",
      "عبد الرحمن الشايب": "6",
      "عمر سند": "6",
      "محمد عودة": "1",
      "أنس أبو الفتوح": "1",
      "عمرو الصايدي": "6",
      "زيد حطاب": "6"
    },
    "مراجعة قريبة": {
      "زياد بالقاسمي": "6",
      "أحمد الدقاق": "6",
      "عبد الرحمن الشايب": "1",
      "عمر سند": "1",
      "محمد عودة": "6",
      "أنس أبو الفتوح": "6",
      "عمرو الصايدي": "1",
      "زيد حطاب": "1"
    }
  }
}

# Example usage
# update_student_records(r"C:\Users\HELR_LPTP\OneDrive\Desktop\class6.xlsx", decoded_data, 'updated_class6.xlsx')