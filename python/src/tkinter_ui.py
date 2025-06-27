import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from openpyxl import load_workbook
import pandas as pd
from encoder_decoder import decode_data
from excel_processor import process_excel_file

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel File Processor")
        
        # Configure the root window to be more spacious
        self.root.geometry("1200x800")
        
        # Create the main frame
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create the folder selection frame
        self.create_folder_selection()
        
        # Create the notebook (tabbed interface) with tabs on the left
        style = ttk.Style()
        style.configure('LeftTab.TNotebook', tabposition='wn')
        
        # Main notebook for files
        self.notebook = ttk.Notebook(self.main_frame, style='LeftTab.TNotebook')
        self.notebook.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        # Dictionaries to store widgets for each file
        self.text_widgets = {}
        self.sheet_notebooks = {}
        self.sheet_tables = {}
        
        # Initially, no folder is selected
        self.current_folder = None
        
    def create_folder_selection(self):
        """Creates the folder selection frame with entry and browse button."""
        folder_frame = ttk.Frame(self.main_frame)
        folder_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Folder path entry
        self.folder_var = tk.StringVar()
        self.folder_entry = ttk.Entry(folder_frame, textvariable=self.folder_var, state='readonly')
        self.folder_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        # Browse button
        browse_btn = ttk.Button(folder_frame, text="Browse", command=self.browse_folder)
        browse_btn.pack(side=tk.RIGHT)
        
    def update_current_file(self, filename):
        """Handler for the Update File button."""
        try:
            # Get the encoded text from the text widget
            encoded_text = self.text_widgets[filename].get("1.0", tk.END).strip()
            if not encoded_text:
                messagebox.showerror("Error", "Please enter the encoded data first.")
                return
                
            # Try to decode the text
            try:
                json_data = decode_data(encoded_text)
            except ValueError as e:
                messagebox.showerror("Decoding Error", f"Failed to decode the data: {str(e)}")
                return
                
            # Get the full path of the current Excel file
            excel_path = os.path.join(self.current_folder, filename)
            
            # Process the Excel file with our decoded data
            try:
                results = process_excel_file(excel_path, json_data)
                
                # Check results and show appropriate message
                failures = [name for name, result in results.items() 
                          if not result['success']]
                
                if failures:
                    error_msg = "Failed to update some students:\n"
                    for name in failures:
                        error_msg += f"- {name}: {results[name]['error']}\n"
                    messagebox.showwarning("Partial Success", error_msg)
                else:
                    messagebox.showinfo("Success", "File updated successfully!")
                    
                # Refresh the current view
                self.refresh_current_tab()
                
            except Exception as e:
                messagebox.showerror("Processing Error", 
                                   f"Failed to process Excel file: {str(e)}")
                
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")
            
    def refresh_current_tab(self):
        """Refreshes the current tab's content after an update."""
        current_tab = self.notebook.select()
        if current_tab:
            tab_id = self.notebook.index(current_tab)
            filename = self.notebook.tab(tab_id)['text']
            
            # Store the current position of the text widget
            text_widget = self.text_widgets.get(filename)
            if text_widget:
                text_content = text_widget.get("1.0", tk.END)
            
            # Reload the tab
            self.create_file_tab(filename)
            
            # Restore the text content
            if text_content:
                self.text_widgets[filename].insert("1.0", text_content)

    def create_file_tab(self, filename):
        """Creates a new tab for an Excel file with text area and sheet tabs."""
        # Create a frame for the file tab
        tab_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab_frame, text=filename)
        
        # Make the tab frame expandable
        tab_frame.grid_rowconfigure(1, weight=1)
        tab_frame.grid_columnconfigure(0, weight=1)
        
        # Create top frame for text area and update button
        top_frame = ttk.Frame(tab_frame)
        top_frame.grid(row=0, column=0, sticky='nsew', pady=(0, 10))
        top_frame.columnconfigure(0, weight=1)  # Make text area expandable
        
        # Create text widget frame (for text and scrollbar)
        text_frame = ttk.Frame(top_frame)
        text_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 10))
        
        # Create and configure text widget
        text_widget = tk.Text(text_frame, wrap=tk.WORD, height=10)
        text_widget.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        
        # Add scrollbar for text area
        text_scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text_widget.config(yscrollcommand=text_scrollbar.set)
        
        # Create Update File button with custom style
        style = ttk.Style()
        style.configure('Update.TButton', padding=(20, 10))
        update_btn = ttk.Button(
            top_frame,
            text="Update File",
            command=lambda: self.update_current_file(filename),
            style='Update.TButton'
        )
        update_btn.grid(row=0, column=1, sticky='ns')
        
        # Store the text widget reference
        self.text_widgets[filename] = text_widget
        
        # Create notebook for sheets (tabs on the left)
        sheet_notebook = ttk.Notebook(tab_frame, style='LeftTab.TNotebook')
        sheet_notebook.grid(row=1, column=0, sticky='nsew')
        
        # Store the sheet notebook reference
        self.sheet_notebooks[filename] = sheet_notebook
        
        # Load and display the Excel file
        try:
            filepath = os.path.join(self.current_folder, filename)
            workbook = load_workbook(filepath, read_only=True, data_only=True)
            
            self.sheet_tables[filename] = {}
            
            for sheet_name in workbook.sheetnames:
                df = pd.read_excel(filepath, sheet_name=sheet_name)
                
                sheet_frame = ttk.Frame(sheet_notebook)
                sheet_notebook.add(sheet_frame, text=sheet_name)
                
                table_frame = self.create_table_frame(sheet_frame, df)
                table_frame.pack(fill=tk.BOTH, expand=True)
                
            workbook.close()
            
        except Exception as e:
            error_frame = ttk.Frame(sheet_notebook)
            sheet_notebook.add(error_frame, text="Error")
            error_label = ttk.Label(
                error_frame,
                text=f"Error loading Excel file:\n{str(e)}",
                foreground="red",
                padding="20"
            )
            error_label.pack(expand=True)

    def create_table_frame(self, parent, data):
        """Creates a scrollable frame containing a table."""
        frame = ttk.Frame(parent)
        
        y_scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL)
        x_scrollbar = ttk.Scrollbar(frame, orient=tk.HORIZONTAL)
        
        table = ttk.Treeview(frame, yscrollcommand=y_scrollbar.set, 
                           xscrollcommand=x_scrollbar.set)
        
        y_scrollbar.config(command=table.yview)
        x_scrollbar.config(command=table.xview)
        
        table.grid(row=0, column=0, sticky='nsew')
        y_scrollbar.grid(row=0, column=1, sticky='ns')
        x_scrollbar.grid(row=1, column=0, sticky='ew')
        
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        
        table['columns'] = list(data.columns)
        table.column('#0', width=0, stretch=tk.NO)
        
        for col in data.columns:
            table.heading(col, text=col)
            table.column(col, width=100)
        
        for idx, row in data.iterrows():
            table.insert('', 'end', values=list(row))
        
        return frame

    def browse_folder(self):
        """Handles folder selection and updates the UI accordingly."""
        folder = filedialog.askdirectory()
        if folder:
            self.current_folder = folder
            self.folder_var.set(folder)
            self.update_tabs()
            
    def update_tabs(self):
        """Updates the tabs based on Excel files in the selected folder."""
        for tab in self.notebook.tabs():
            self.notebook.forget(tab)
        self.text_widgets.clear()
        self.sheet_notebooks.clear()
        self.sheet_tables.clear()
        
        excel_files = [f for f in os.listdir(self.current_folder) 
                      if f.endswith(('.xlsx', '.xls'))]
        
        if excel_files:
            for file in excel_files:
                self.create_file_tab(file)
        else:
            message_frame = ttk.Frame(self.notebook)
            self.notebook.add(message_frame, text="No Files")
            message_label = ttk.Label(
                message_frame, 
                text="No Excel files found in the selected folder",
                padding="20"
            )
            message_label.pack(expand=True)

def main():
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()